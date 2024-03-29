import investpy
import xlrd
import pandas as pd
from openpyxl import load_workbook, Workbook
from progress.bar import IncrementalBar

SHORT_NAMES_ROW_NUM = 2  # Vital table column positions
COUNTRIES_ROW_NUM = 3
SHEET_NAME = 'SP500'

def check_output_file():  # Checking if there's a local source and output files
    try:
        with open('data.txt', 'r') as outputtxt:
            pathsfromfile = [outputtxt.readline()[:-1], outputtxt.readline()]
        return pathsfromfile
    except:
        return False

def no_data_found_message(stock):
    return [[stock]+['NO DATA']*5]

def main(supervisor):  # Supervisor is a progressbar UI view
    paths = check_output_file()  # Getting paths to necessary files
    path_to_file = paths[0]
    path_to_output = paths[1]

    with xlrd.open_workbook(path_to_file) as book:  # Getting local stocks names for loop
        sheet = book.sheet_by_name(SHEET_NAME)
    
    cell_iterator = sheet.nrows  # Number of none-empty cells

    data = [sheet.cell_value(r, SHORT_NAMES_ROW_NUM) for r in range(1, cell_iterator)]  # An array of stock short names
    country_data = [sheet.cell_value(r, COUNTRIES_ROW_NUM) for r in range(1, cell_iterator)]  # Country list for multinational stock lists

    wb = Workbook()  # Creating workbook for results
    wb.save(path_to_output)
    
    book = load_workbook(path_to_output)  # Working with output file without overwriting it within loop
    writer = pd.ExcelWriter(path_to_output, engine='openpyxl')

    i = 0  # Loop iterators
    countryiterator = 0
    queue_len = len(data)

    bar = IncrementalBar('Processing', max=queue_len)  # Initializing the progress bar

    supervisor.max_value = queue_len
    supervisor.progress.setMaximum(queue_len)  # UI progressbar setting

    for stock in data:  # Loop getting dividends by stock name using investpy funcs
        try:
            stock_info = investpy.stocks.get_stock_dividends(stock, country=country_data[countryiterator])
            
            stock_info.insert(0, "Name", stock, True)
            stock_info = stock_info.iloc[:8]  # Slicing data for a period we need
            if i != 0:
                stock_info.to_excel(writer, "Sheet", startrow=i, header=False, index=False)
                writer.save()
                i += 9
            else:
                stock_info.to_excel(writer, "Sheet", startrow=i, header=True, index=False)
                writer.save()
                i += 10
            bar.next()
            print(' - ', stock, ' - DATA FOUND')
        except RuntimeError:  # If there is no data provided - write 'NO DATA' to every info cell
            df = pd.DataFrame(no_data_found_message(stock))
            df.to_excel(writer, "Sheet", startrow=i, header=False, index=False)
            writer.save()
            i += 2
            bar.next()
            print(' - ', stock, ' - DATA NOT FOUND - RuntimeError occured')
        except ValueError:
            df = pd.DataFrame(no_data_found_message(stock))
            df.to_excel(writer, "Sheet", startrow=i, header=False, index=False)
            writer.save()
            i += 2
            bar.next()
            print(' - ', stock, ' - DATA NOT FOUND - ValueError occured')
        except KeyboardInterrupt:  # Wrapped keyboard interruption traceback
            print("\n#-#-# Interrupted by User #-#-#")
            exit(0)
        finally:
            supervisor.progress_value += 1
            supervisor.progress.setValue(supervisor.progress_value)
            supervisor.info_label.setText('Fetching data: ' + str(supervisor.progress_value) + ' / ' + str(supervisor.max_value))
            countryiterator += 1
    bar.finish()
    supervisor.on_stop()
